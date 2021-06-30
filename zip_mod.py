# -*- coding: utf-8 -*-
"""
Created on Wed Jun 23 16:25:15 2021

@author: thomas.collaudin
"""

from zipfile import ZipFile
from os.path import basename, isfile, join
from os import listdir
import os
import openpyxl
from datetime import datetime

def zip_folder(zip_path, folder_path, delete_files = False) :
    zipped = ZipFile(zip_path, 'w')
    files = list_files(folder_path)
    for file in files :
        zipped.write(file, basename(file))
        if delete_files == True :
            os.remove(file)
    zipped.close()
    return zipped
    
def list_files(folder_path) :
    return [folder_path + f for f in listdir(folder_path) if isfile(join(folder_path, f))]

def histo_results(wb_path, img_name, now_str, nb_contour, nb_model, nb_combined) :
    wb = openpyxl.load_workbook(filename = wb_path)
    ws = wb['histo']
    last_r = ws.max_row
    last_c = ws.max_column
    for c in range(2, last_c + 3) :
        if ws.cell(2, c).value == None :
            last_c = c
            break
    myc = last_c
    for c in range(2, last_c + 3) :
        if ws.cell(1, c).value == now_str :
            myc = c
            break
    ws.cell(1, myc).value = now_str
    ws.cell(2, myc).value = 'Contour'
    ws.cell(2, myc + 1).value = 'Model'
    ws.cell(2, myc + 2).value = 'Combined'
    for r in range(3, last_r + 3) :
        if ws.cell(r, 1).value == img_name :
            ws.cell(r, myc).value = nb_contour
            ws.cell(r, myc + 1).value = nb_model
            ws.cell(r, myc + 2).value = nb_combined
            break
        elif ws.cell(r, 1).value == None :
            ws.cell(r, 1).value = img_name
            ws.cell(r, myc).value = nb_contour
            ws.cell(r, myc + 1).value = nb_model
            ws.cell(r, myc + 2).value = nb_combined
            break
    wb.save(wb_path)
    
# wb = openpyxl.load_workbook(filename = 'results/results.xlsx')
# ws = wb['histo']
# ws.cell(1,10).value = 'blabla'
# wb.save('results/results.xlsx')

# histo_results('results/results.xlsx', 'img_nam', 42, 'com')

zip_path = 'boxes_1.zip'
folder_path = 'results/boxes/'

